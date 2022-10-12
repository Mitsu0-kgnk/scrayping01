import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
# import chromedriver_binary
from time import sleep
import openpyxl
from openpyxl import Workbook, load_workbook
import schedule


chrome_path = '/Users/kuginukimitsuruhiroshi/Desktop/Portfolio/Scrayping/driver/chromedriver.exe'
options = Options()
options.add_argument('--headless')

browser = webdriver.Chrome(options=options)

url = 'https://daidata.goraggio.com/store_list?pref=%E5%8D%83%E8%91%89%E7%9C%8C'


browser.get(url)

sleep(2)
stores = browser.find_elements_by_tag_name('tr')
store = stores[5]

store.click()

sleep(2)

accept_btn = browser.find_element_by_class_name('accept_btn')
form = accept_btn.find_element_by_tag_name('form')
form.submit()

sleep(2)

menu = browser.find_elements_by_class_name('Radius-Basic')
menu[5].click()

sleep(2)

data = []

numbers = browser.find_elements_by_class_name('Text-UnderLine')
num = len(numbers)-1

a = browser.find_elements_by_class_name('Text-UnderLine')
a = a[:-1]
links = [i.get_attribute('href') for i in a]
for link in links:
    sleep(2)

    browser.get(link)

    h4 = browser.find_elements_by_tag_name('h4')
    date = h4[0].text
    date = date.split(' ')[0]


    h1 = browser.find_element_by_id('pachinkoTi')
    machine_name = h1.find_element_by_tag_name('strong').text 
    sub = h1.find_element_by_tag_name('span').text
    sub = sub.replace('(','').replace(')','')
    sub = sub.split('|')
    category = sub[0]
    number = sub[1]

    l = ['日付',
            '名前',
            '種類',
            '台番号',
            '大当たり',
            '確変',
            'スタート',
            '大当たり確率',
            '最大持玉',
            '累計スタート',
            '初当たり確率',
            '前日最終スタート'
        ]

    cotents = [date,machine_name,category,number]


    tables_2 = browser.find_element_by_class_name('overviewTable2')
    table_2 = tables_2

    scores = table_2.find_elements_by_tag_name('td')

    score1 = []

    for item in scores:
        score1.append(item.text)


    tables_3 = browser.find_elements_by_class_name('overviewTable3')
    table_3 = tables_3[1]
    trs = table_3.find_elements_by_tag_name('tr')
    tr = trs[0]


    score2 = []

    td = tr.find_elements_by_tag_name('td')

    for i in trs:
        td = i.find_elements_by_tag_name('td')
        for item in td:
            score2.append(item.text)
            
    try:
        score1[3] = eval(score1[3])
        score2[2] = eval(score2[2])
    except ZeroDivisionError:
        score1[3] = float(0.0)
        score2[2] = float(0.0)
    items = cotents + score1 + score2

    d = dict(zip(l,items))

    data.append(d)


    browser.back()

df = pd.DataFrame(data)

wb = openpyxl.load_workbook('PA_DATA10.xlsx')
sheets = wb.sheetnames

for i,sheet in enumerate(sheets):
    ws = wb[sheet]
    a = list(data[i].values())
    ws.append(a)

wb.save('PA_DATA10.xlsx') 